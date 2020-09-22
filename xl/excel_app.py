import openpyxl as xl
from openpyxl.chart import Reference, PieChart3D, BarChart3D
from pathlib import Path
import os


def file_updater(folder):
    print("Fetching files ... \nPlease wait . . .")
    directory = f"/mnt/c/Users/7K/{str(folder.strip())}"
    filename = Path()
    os.chdir(directory)
    print(f"Currently working at: '{directory}' directory.")
    n = 0
    for i, filename in enumerate(filename.glob("*.xlsx")):
        wb = xl.load_workbook(filename)
        sheet = wb["Sheet1"]
        cell_ttl = sheet.cell(1, sheet.max_column + 1, value="Updated Price")
        # cell_ttl.value = "Updated Price"
        for row in range(2, sheet.max_row + 1):
            old_cell = sheet.cell(row, sheet.max_column - 1)
            upd_val = (old_cell.value) * (1.25)
            new_cell = sheet.cell(row, sheet.max_column)
            new_cell.value = upd_val

        values = Reference(
            worksheet=sheet, min_col=4, max_col=4, min_row=2, max_row=sheet.max_row
        )

        chart1, chart2 = BarChart3D(), PieChart3D()
        chart1.add_data(values)
        chart2.add_data(values)

        place1, place2 = "f2", "f20"
        sheet.add_chart(chart1, place1)
        sheet.add_chart(chart2, place2)

        wb.save(f"New_{i+1}-{filename}")
        n += 1
    print(f"({n} Files Found and Updated)")
    return print("\nFiles Succesfully Updated.")


data = input("Type the directory you want to update Excel('*.xlsx') files:  ").strip()
# data = "Documents/Python/master-python101/xl"
make = file_updater(data)
