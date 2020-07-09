import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference
from pathlib import Path


def price_updater(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    # cell = sheet["a1"]
    cell = sheet.cell(1, 1)
    print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        print(row)
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = (cell.value) * (0.9)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    cell_updated = sheet.cell(1, 4)
    cell_updated.value = "Updated"

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(filename)


filename = Path()
for filename in filename.glob("*.xlsx"):
    price_updater(filename)
